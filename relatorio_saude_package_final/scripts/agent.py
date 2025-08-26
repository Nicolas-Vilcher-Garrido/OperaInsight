#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
agent.py
==========

This module implements an Excel reporting agent for health data. The agent reads
CSV/TXT files produced by SigSaúde and Operadoras, cleans and merges them,
calculates key metrics, generates pivots and rankings, and writes the results
into a macro-enabled Excel workbook. It also records an audit trail and logs
every run. The aim is to automate the repetitive steps involved in preparing
management reports from raw transactional data.

Key features implemented:

* **Auto-discovery of input files**: scans the configured directories for
  `.csv` and `.txt` files.
* **Flexible CSV/TXT import**: detects delimiters and tries common encodings
  (UTF‑8 and Latin1) to load files robustly.
* **Data cleaning**: converts dates, coerces numerics, replaces negatives with
  zero, computes missing revenue values, removes duplicates and filters by
  period/UF/Categoria.
* **VLOOKUP/merge**: enriches the dataset with a Segmento column by joining
  against a DimClientes table. The merge is performed in Python but a VLOOKUP
  formula is also inserted in the Excel sheet for demonstration purposes.
* **Pivot and rankings**: summarises revenue by month, category and operadora,
  and produces top lists for operadoras and procedimentos.
* **Excel output**: writes multiple sheets (Dados, Resumo, DimClientes,
  Rankings, Parametros, Auditoria) with proper formatting, formulas and
  conditional formatting. Generates example charts (line, column and pie) on
  the Resumo sheet.
* **Logging and auditing**: records a detailed audit in both an Excel sheet
  and a rotating log file.

The main entry points are:

```
atualizar_tudo()  # runs a full end‑to‑end update
gerar_graficos()  # regenerates charts on an existing workbook
gerar_rankings()  # updates rankings only
```

The macros contained in ``vba/ModuleExcelAgent.bas`` should call these
functions via Shell or xlwings to expose them to end users.

Note: macros are not embedded automatically into the workbook by this script.
They are provided as a separate `.bas` file and must be imported into the
workbook manually or via an Excel add‑in. This limitation arises because the
environment does not support generating a compiled VBA project on the fly.
"""

from __future__ import annotations

import csv
import logging
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import pandas as pd
import numpy as np
import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Color
from openpyxl.chart import LineChart, Reference, Series, BarChart, PieChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Configuration loading and parameter management
# ---------------------------------------------------------------------------


def load_config(config_path: Optional[str] = None) -> Dict:
    """Load configuration from YAML. Defaults to ``config.yaml`` in the parent
    folder of this script.

    Args:
        config_path: Optional explicit path to config file.

    Returns:
        Parsed configuration dictionary.
    """
    script_dir = Path(__file__).resolve().parent
    default_path = script_dir.parent / "config.yaml"
    cfg_path = Path(config_path) if config_path else default_path
    if not cfg_path.exists():
        raise FileNotFoundError(f"Config file not found: {cfg_path}")
    with open(cfg_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    return config


def read_params_from_sheet_or_cfg(cfg: Dict) -> Dict:
    """Read parameters (period, UFs, categories) from the Parametros sheet of
    an existing workbook. If the workbook or sheet doesn't exist, fall back
    to the values defined in the config.

    The Parametros sheet layout is assumed to match the specification:
    cell C2 = periodo_inicio
    cell C3 = periodo_fim
    cell C4 = comma‑separated list of UFs
    cell C5 = comma‑separated list of Categorias

    Args:
        cfg: Configuration dictionary.

    Returns:
        Dict with keys: periodo_inicio, periodo_fim, ufs (list), categorias (list).
    """
    params = {
        "periodo_inicio": cfg["parametros"]["periodo_inicio"],
        "periodo_fim": cfg["parametros"]["periodo_fim"],
        "ufs": cfg["parametros"]["uf_incluir"],
        "categorias": cfg["parametros"]["categoria_incluir"],
    }
    workbook_path = Path(cfg["saida"]["relatorio_xlsm"])
    if not workbook_path.exists():
        return params
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        if "Parametros" not in wb.sheetnames:
            return params
        ws = wb["Parametros"]
        # Excel is 1‑indexed; 'C' is column 3
        periodo_inicio_val = ws["C2"].value
        periodo_fim_val = ws["C3"].value
        ufs_val = ws["C4"].value
        categorias_val = ws["C5"].value
        if periodo_inicio_val:
            params["periodo_inicio"] = str(periodo_inicio_val)
        if periodo_fim_val:
            params["periodo_fim"] = str(periodo_fim_val)
        if ufs_val:
            if isinstance(ufs_val, str):
                params["ufs"] = [u.strip() for u in ufs_val.split(",") if u.strip()]
        if categorias_val:
            if isinstance(categorias_val, str):
                params["categorias"] = [c.strip() for c in categorias_val.split(",") if c.strip()]
    except Exception as exc:
        # On any failure, fall back to config
        logging.getLogger(__name__).warning(
            f"Failed to read Parametros sheet from {workbook_path}: {exc}. Using defaults."
        )
    return params


# ---------------------------------------------------------------------------
# Data import utilities
# ---------------------------------------------------------------------------


def autodiscover_files(entrada_cfg: Dict) -> List[str]:
    """Discover CSV and TXT files under the configured input directories.

    Args:
        entrada_cfg: Dictionary with keys ``sigsaude_dir`` and ``operadoras_dir``.

    Returns:
        List of absolute file paths to process.
    """
    files: List[str] = []
    project_root = Path(__file__).resolve().parents[1]
    for key in ("sigsaude_dir", "operadoras_dir"):
        base = Path(entrada_cfg[key])
        if not base.is_absolute():
            base = (project_root / base).resolve()
        if not base.exists():
            continue
        for path in base.rglob("*"):
            if path.suffix.lower() in {".csv", ".txt"} and path.is_file():
                files.append(str(path))
    return files


def _detect_separator(line: str) -> str:
    """Heuristically detect the delimiter used in a CSV/TXT header line.
    Considers common separators and returns the one that yields the most fields.
    """
    candidates = [",", ";", "\t", "|"]
    best_sep = ","
    best_count = 0
    for sep in candidates:
        count = line.count(sep)
        if count > best_count:
            best_count = count
            best_sep = sep
    return best_sep


def _read_single_file(path: str) -> pd.DataFrame:
    """Read a single CSV/TXT file, trying to detect delimiter and encoding.

    Args:
        path: Path to the file.

    Returns:
        DataFrame with the contents. If loading fails, returns an empty DataFrame.
    """
    try:
        with open(path, "rb") as f:
            sample = f.read(2048)
        # Try UTF‑8 first
        encoding = "utf-8"
        try:
            sample.decode(encoding)
        except UnicodeDecodeError:
            encoding = "latin1"
        # Detect separator from first non‑empty line in decoded sample
        decoded_sample = sample.decode(encoding, errors="ignore")
        first_line = next((line for line in decoded_sample.splitlines() if line.strip()), "")
        sep = _detect_separator(first_line)
        df = pd.read_csv(path, sep=sep, encoding=encoding, engine="python")
        return df
    except Exception as exc:
        logging.getLogger(__name__).error(f"Failed to read {path}: {exc}")
        return pd.DataFrame()


def read_and_concat(files: List[str]) -> pd.DataFrame:
    """Load and concatenate multiple CSV/TXT files.

    Column names are normalised to match the expected names (case insensitive).

    Args:
        files: List of file paths to read.

    Returns:
        Concatenated DataFrame.
    """
    data_frames = []
    for file_path in files:
        df = _read_single_file(file_path)
        if df.empty:
            continue
        # Normalise columns: strip whitespace and title‑case to match spec
        df.columns = [str(c).strip() for c in df.columns]
        # Rename columns ignoring case differences
        rename_map = {}
        for col in df.columns:
            lower = col.lower()
            if lower == "data":
                rename_map[col] = "Data"
            elif lower == "clienteid" or lower == "cliente_id" or lower == "cliente id":
                rename_map[col] = "ClienteId"
            elif lower == "uf":
                rename_map[col] = "UF"
            elif lower == "operadora":
                rename_map[col] = "Operadora"
            elif lower == "procedimento":
                rename_map[col] = "Procedimento"
            elif lower == "categoria":
                rename_map[col] = "Categoria"
            elif lower == "qtde" or lower == "quantidade":
                rename_map[col] = "Qtde"
            elif lower in ("preco", "preco_unitario", "preçounitario", "precounitario"):
                rename_map[col] = "PrecoUnitario"
            elif lower == "receita":
                rename_map[col] = "Receita"
        df = df.rename(columns=rename_map)
        data_frames.append(df)
    if not data_frames:
        return pd.DataFrame()
    return pd.concat(data_frames, ignore_index=True, sort=False)


# ---------------------------------------------------------------------------
# Data cleansing and transformation
# ---------------------------------------------------------------------------


def clean_data(df: pd.DataFrame, params: Dict) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Clean the raw dataset according to business rules.

    Steps performed:
    * Coerce date column to datetime and filter outside the period.
    * Coerce numeric columns and replace negatives with zero.
    * Compute missing 'Receita' as Qtde * PrecoUnitario.
    * Remove duplicates on (Data, ClienteId, Procedimento).
    * Apply UF and Categoria filters.

    Args:
        df: Raw concatenated DataFrame.
        params: Dict containing periodo_inicio, periodo_fim, ufs, categorias.

    Returns:
        Tuple of (cleaned DataFrame, counts dict with details of dropped records).
    """
    counts = {
        "total_lido": len(df),
        "negativos_corrigidos": 0,
        "duplicados_removidos": 0,
        "fora_periodo_removidos": 0,
        "filtro_uf_categoria_removidos": 0,
    }
    if df.empty:
        return df, counts
    # Ensure expected columns exist; if not, create them with NaNs
    for col in ["Data", "ClienteId", "UF", "Operadora", "Procedimento", "Categoria", "Qtde", "PrecoUnitario"]:
        if col not in df.columns:
            df[col] = np.nan
    # Convert Data to datetime
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    # Coerce numerics
    for col in ["Qtde", "PrecoUnitario", "Receita"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    # Replace NaN in Qtde/PrecoUnitario with 0
    df["Qtde"] = df["Qtde"].fillna(0)
    df["PrecoUnitario"] = df["PrecoUnitario"].fillna(0)
    # Replace negatives with 0 and count occurrences
    for col in ["Qtde", "PrecoUnitario"]:
        neg_mask = df[col] < 0
        if neg_mask.any():
            counts["negativos_corrigidos"] += int(neg_mask.sum())
            df.loc[neg_mask, col] = 0
    # Compute Receita if missing or NaN
    if "Receita" not in df.columns:
        df["Receita"] = df["Qtde"] * df["PrecoUnitario"]
    else:
        df["Receita"] = df["Receita"].fillna(df["Qtde"] * df["PrecoUnitario"])
    # Drop duplicates based on key subset
    before_dups = len(df)
    df = df.drop_duplicates(subset=["Data", "ClienteId", "Procedimento"], keep="first")
    counts["duplicados_removidos"] = before_dups - len(df)
    # Filter by date
    inicio = pd.to_datetime(params["periodo_inicio"], errors="coerce")
    fim = pd.to_datetime(params["periodo_fim"], errors="coerce")
    if not pd.isna(inicio) and not pd.isna(fim):
        mask_period = df["Data"].between(inicio, fim)
        counts["fora_periodo_removidos"] = int((~mask_period).sum())
        df = df[mask_period]
    # Filter by UF and Categoria
    ufs = params.get("ufs") or []
    categorias = params.get("categorias") or []
    mask_filters = df["UF"].isin(ufs) & df["Categoria"].isin(categorias)
    counts["filtro_uf_categoria_removidos"] = int((~mask_filters).sum())
    df = df[mask_filters]
    # Reset index after filtering
    df = df.reset_index(drop=True)
    return df, counts


def vlookup_segment(df: pd.DataFrame, cfg: Dict) -> Tuple[pd.DataFrame, int]:
    """Perform a merge between the main dataset and the DimClientes table to
    enrich with the Segmento column.

    Args:
        df: Cleaned DataFrame.
        cfg: Configuration dict.

    Returns:
        Tuple of (DataFrame with Segmento column, number of rows where segment
        was missing).
    """
    dim_path = Path(cfg["chaves"]["dim_clientes"])
    # Resolve relative paths relative to the project root (two levels up from this file)
    if not dim_path.is_absolute():
        # The config paths are relative to the workbook root. Build absolute path accordingly.
        project_root = Path(__file__).resolve().parents[1]
        dim_path = (project_root / dim_path).resolve()
    if not dim_path.exists():
        logging.getLogger(__name__).warning(f"DimClientes file missing: {dim_path}")
        df["Segmento"] = np.nan
        return df, len(df)
    dim_df = pd.read_csv(dim_path)
    # Normalise columns
    dim_df.columns = [c.strip() for c in dim_df.columns]
    # If the main DF does not have the key column, return with NaNs
    if "ClienteId" not in df.columns:
        df["Segmento"] = np.nan
        return df, len(df)
    # Identify join columns
    cliente_key_cfg = cfg["chaves"].get("cliente_id", "ClienteId")
    dim_key = None
    for col in dim_df.columns:
        if col.lower() == cliente_key_cfg.lower():
            dim_key = col
            break
    if dim_key is None:
        dim_key = dim_df.columns[0]
    # Ensure 'Segmento' exists in dim_df
    if "Segmento" not in dim_df.columns:
        dim_df["Segmento"] = np.nan
    df = df.merge(dim_df[[dim_key, "Segmento"]], how="left", left_on="ClienteId", right_on=dim_key)
    if dim_key != "ClienteId":
        df.drop(columns=[dim_key], inplace=True)
    missing_segments = df["Segmento"].isna().sum()
    return df, int(missing_segments)


def pivot_data(df: pd.DataFrame) -> pd.DataFrame:
    """Create a pivot table summarising revenue by month, category and operadora.

    Args:
        df: Cleaned DataFrame.

    Returns:
        Pivoted DataFrame with multi‑level columns.
    """
    if df.empty:
        return pd.DataFrame()
    # Derive month column
    df["Mes"] = df["Data"].dt.to_period("M").dt.to_timestamp()
    pv = pd.pivot_table(
        df,
        values="Receita",
        index="Mes",
        columns=["Categoria", "Operadora"],
        aggfunc="sum",
        fill_value=0,
    )
    pv = pv.sort_index().reset_index()
    return pv


def generate_rankings(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Generate rankings of operadoras and procedimentos by total revenue.

    Args:
        df: Cleaned DataFrame.

    Returns:
        Tuple of (operadora ranking DF, procedimento ranking DF).
    """
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()
    oper = df.groupby("Operadora")["Receita"].sum().reset_index()
    oper = oper.sort_values("Receita", ascending=False).reset_index(drop=True)
    oper.index += 1  # rank starting at 1
    oper.rename(columns={"Receita": "TotalReceita"}, inplace=True)
    oper["Rank"] = oper.index
    proc = df.groupby("Procedimento")["Receita"].sum().reset_index()
    proc = proc.sort_values("Receita", ascending=False).reset_index(drop=True)
    proc.index += 1
    proc.rename(columns={"Receita": "TotalReceita"}, inplace=True)
    proc["Rank"] = proc.index
    return oper, proc


# ---------------------------------------------------------------------------
# Excel output generation
# ---------------------------------------------------------------------------


def write_excel_sheets(
    df: pd.DataFrame,
    pivot_df: pd.DataFrame,
    oper_rank: pd.DataFrame,
    proc_rank: pd.DataFrame,
    params: Dict,
    cfg: Dict,
    counts: Dict[str, int],
    missing_segments: int,
) -> None:
    """Write the multiple sheets to an Excel workbook (.xlsm).

    This function uses pandas to output DataFrames and then openpyxl to apply
    formulas, charts and formatting.
    """
    out_path = Path(cfg["saida"]["relatorio_xlsm"])
    # Resolve relative output path to the project root
    if not out_path.is_absolute():
        project_root = Path(__file__).resolve().parents[1]
        out_path = (project_root / out_path).resolve()
    # Ensure parent dir exists
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Write initial sheets using pandas
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)
        # Flatten MultiIndex columns for pivot, if present
        flat_pivot = pivot_df.copy()
        if isinstance(flat_pivot.columns, pd.MultiIndex):
            new_cols = []
            for col in flat_pivot.columns:
                if isinstance(col, tuple):
                    # Combine levels, ignoring None
                    parts = [str(c) for c in col if c not in ["", None]]
                    new_cols.append("_".join(parts))
                else:
                    new_cols.append(str(col))
            flat_pivot.columns = new_cols
        flat_pivot.to_excel(writer, sheet_name="Resumo", index=False)
        # Rankings: write both rankings into the same sheet at different positions.
        oper_rank.to_excel(writer, sheet_name="Rankings", startrow=1, index=False)
        # Write procedimento ranking below oper ranking with some spacing. We'll re‑apply titles later.
        proc_start = len(oper_rank) + 4
        proc_rank.to_excel(writer, sheet_name="Rankings", startrow=proc_start, index=False)
        # DimClientes sheet: copy raw CSV for reference
        dim_path = Path(cfg["chaves"]["dim_clientes"])
        if dim_path.exists():
            dim_df = pd.read_csv(dim_path)
            dim_df.to_excel(writer, sheet_name="DimClientes", index=False)
        # Parametros sheet: write values into cells (no DataFrame)
        # We'll leave it blank here; openpyxl will populate later
        wb = writer.book
        if "Parametros" not in wb.sheetnames:
            wb.create_sheet("Parametros")
        if "Auditoria" not in wb.sheetnames:
            wb.create_sheet("Auditoria")
    # Now reopen with openpyxl to apply formulas and formatting
    wb = load_workbook(out_path)
    # Ensure all expected sheets exist
    ws_dados = wb["Dados"]
    ws_resumo = wb["Resumo"]
    ws_rank = wb["Rankings"]
    ws_param = wb["Parametros"]
    ws_auditoria = wb["Auditoria"]
    # ------------------------------------------------------------------
    # Parametros sheet
    ws_param["B1"] = "Parâmetro"
    ws_param["C1"] = "Valor"
    ws_param["B2"] = "Periodo início"
    ws_param["C2"] = params["periodo_inicio"]
    ws_param["B3"] = "Periodo fim"
    ws_param["C3"] = params["periodo_fim"]
    ws_param["B4"] = "UFs"
    ws_param["C4"] = ",".join(params["ufs"])
    ws_param["B5"] = "Categorias"
    ws_param["C5"] = ",".join(params["categorias"])
    # Apply bold to header
    bold_font = Font(bold=True)
    for cell in [ws_param["B1"], ws_param["C1"]]:
        cell.font = bold_font
    # ------------------------------------------------------------------
    # Auditoria sheet: log counts
    audit_headers = [
        "DataHoraExecucao",
        "PeriodoInicio",
        "PeriodoFim",
        "TotalLido",
        "NegativosCorrigidos",
        "DuplicadosRemovidos",
        "ForaPeriodoRemovidos",
        "FiltroUF_CategoriaRemovidos",
        "TotalFinal",
        "RegistrosSemSegmento",
    ]
    audit_values = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        params["periodo_inicio"],
        params["periodo_fim"],
        counts["total_lido"],
        counts["negativos_corrigidos"],
        counts["duplicados_removidos"],
        counts["fora_periodo_removidos"],
        counts["filtro_uf_categoria_removidos"],
        len(df),
        missing_segments,
    ]
    for col_idx, header in enumerate(audit_headers, start=1):
        ws_auditoria.cell(row=1, column=col_idx, value=header).font = bold_font
        ws_auditoria.cell(row=2, column=col_idx, value=audit_values[col_idx - 1])
    # ------------------------------------------------------------------
    # Dados sheet formulas and formatting
    # Determine last column and row
    max_row = ws_dados.max_row
    max_col = ws_dados.max_column
    # Identify indices of key columns
    header = [cell.value for cell in ws_dados[1]]
    try:
        col_idx_cliente = header.index("ClienteId") + 1
        col_idx_receita = header.index("Receita") + 1
    except ValueError:
        col_idx_cliente = 2
        col_idx_receita = max_col
    # Insert a new column to the right of Segmento (or append if absent)
    # We'll store Excel VLOOKUP formula here to demonstrate PROCV usage
    new_col_idx = ws_dados.max_column + 1
    ws_dados.insert_cols(new_col_idx)
    ws_dados.cell(row=1, column=new_col_idx, value="Segmento_PROCV").font = bold_font
    # Write formula for each data row
    dim_sheet_name = "DimClientes"
    for row in range(2, max_row + 1):
        # Use absolute column letter for ClienteId (assumed column B after rename)
        cli_cell = ws_dados.cell(row=row, column=col_idx_cliente).coordinate
        formula = f"=VLOOKUP({cli_cell},{dim_sheet_name}!$A:$B,2,FALSE)"
        ws_dados.cell(row=row, column=new_col_idx, value=formula)
    # Sum formula at bottom of Receita column
    total_row = max_row + 2
    # Ensure the label goes into a valid column (>=1)
    label_col = col_idx_receita - 1 if col_idx_receita > 1 else col_idx_receita
    ws_dados.cell(row=total_row, column=label_col, value="Total").font = bold_font
    sum_formula = f"=SUM({ws_dados.cell(row=2, column=col_idx_receita).coordinate}:{ws_dados.cell(row=max_row, column=col_idx_receita).coordinate})"
    ws_dados.cell(row=total_row, column=col_idx_receita, value=sum_formula)
    # SOMASE example: sum for the first UF in params
    if params["ufs"]:
        target_uf = params["ufs"][0]
        # Determine column index for UF
        try:
            col_idx_uf = header.index("UF") + 1
        except ValueError:
            col_idx_uf = 3
        somase_formula = f"=SUMIF({ws_dados.cell(row=2, column=col_idx_uf).column_letter}:{ws_dados.cell(row=max_row, column=col_idx_uf).column_letter},\"{target_uf}\",{ws_dados.cell(row=2, column=col_idx_receita).column_letter}:{ws_dados.cell(row=max_row, column=col_idx_receita).column_letter})"
        # Write SOMASE label and formula for the first UF
        label_col2 = col_idx_receita - 1 if col_idx_receita > 1 else col_idx_receita
        ws_dados.cell(row=total_row + 1, column=label_col2, value=f"SOMASE({target_uf})").font = bold_font
        ws_dados.cell(row=total_row + 1, column=col_idx_receita, value=somase_formula)
    # SE (IF) formula to flag high values (> P90). We'll compute P90 in Resumo
    # Determine column index for AltoValor (append new column)
    high_col_idx = ws_dados.max_column + 1
    ws_dados.cell(row=1, column=high_col_idx, value="AltoValor").font = bold_font
    # We'll place p90 value into Resumo!B2 later
    for row in range(2, max_row + 1):
        receita_cell = ws_dados.cell(row=row, column=col_idx_receita).coordinate
        formula = f"=IF({receita_cell}>Resumo!$B$2,\"ALTO\",\"\")"
        ws_dados.cell(row=row, column=high_col_idx, value=formula)
    # Apply basic formatting: bold header, freeze top row, date and currency formats
    for col in ws_dados.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold_font
    ws_dados.freeze_panes = ws_dados["A2"]
    # Set column widths roughly based on header lengths
    for i, value in enumerate(header, start=1):
        length = max(len(str(value)), 10)
        ws_dados.column_dimensions[ws_dados.cell(row=1, column=i).column_letter].width = length + 2
    # Format date column
    try:
        col_idx_data = header.index("Data") + 1
        for cell in ws_dados.iter_rows(min_row=2, min_col=col_idx_data, max_col=col_idx_data, max_row=max_row):
            cell[0].number_format = "yyyy-mm-dd"
    except ValueError:
        pass
    # Format numeric columns
    currency_fmt = "[R$-pt-BR] #,##0.00"
    for col_name in ["Qtde", "PrecoUnitario", "Receita", "TotalReceita"]:
        if col_name in header:
            idx = header.index(col_name) + 1
            for cell in ws_dados.iter_rows(min_row=2, min_col=idx, max_col=idx, max_row=max_row + 1):
                # Use integer format for Qtde
                if col_name == "Qtde":
                    cell[0].number_format = "0"
                else:
                    cell[0].number_format = currency_fmt
    # Conditional formatting for Receita column (gradient)
    # Determine min and max for Receita
    receita_values = [row[0].value for row in ws_dados.iter_rows(min_row=2, min_col=col_idx_receita, max_col=col_idx_receita, max_row=max_row)]
    if receita_values:
        try:
            float_values = [float(val) for val in receita_values if val is not None]
            if float_values:
                # Apply 3‑color scale
                rule = ColorScaleRule(
                    start_type="min",
                    start_color="FFFFFF",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFD966",
                    end_type="max",
                    end_color="00B050",
                )
                range_str = f"{ws_dados.cell(row=2, column=col_idx_receita).coordinate}:{ws_dados.cell(row=max_row, column=col_idx_receita).coordinate}"
                ws_dados.conditional_formatting.add(range_str, rule)
        except Exception:
            pass
    # ------------------------------------------------------------------
    # Resumo sheet: place P90 value and generate charts
    # Compute p90 using pandas for accuracy
    if not df.empty:
        p90_value = float(np.percentile(df["Receita"], 90)) if len(df) > 0 else 0.0
    else:
        p90_value = 0.0
    ws_resumo["A1"] = "Mês"
    # Write pivot headers and values are already there; adjust p90 cell
    ws_resumo["B2"] = p90_value
    ws_resumo["B1"] = "P90_Receita"
    ws_resumo["A1"].font = bold_font
    ws_resumo["B1"].font = bold_font
    ws_resumo["B2"].number_format = currency_fmt
    # Generate charts on Resumo
    _add_charts_to_resumo(ws_resumo, df, params, currency_fmt)
    # ------------------------------------------------------------------
    # Rankings sheet formatting: add titles
    ws_rank.cell(row=1, column=1, value="Ranking de Operadoras").font = bold_font
    start_proc = len(oper_rank) + 4
    ws_rank.cell(row=start_proc, column=1, value="Ranking de Procedimentos").font = bold_font
    # Auto column widths in Rankings
    for col in ws_rank.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_rank.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    # Save workbook
    wb.save(out_path)


from openpyxl.worksheet.worksheet import Worksheet


def _add_charts_to_resumo(ws: Worksheet, df: pd.DataFrame, params: Dict, currency_fmt: str) -> None:
    """Internal helper to add line, bar and pie charts to the Resumo sheet.

    Charts are inserted below the pivot table starting at row 4. They
    summarise the overall revenue by month, by operadora and by
    procedimento.
    """
    from openpyxl.utils import get_column_letter
    # Clear any existing charts
    ws._charts = []  # type: ignore
    # Determine where pivot table ends (assuming it starts at row 2 and columns begin at A)
    max_row = ws.max_row
    max_col = ws.max_column
    chart_start_row = max_row + 2
    # Monthly total revenue line chart
    if not df.empty:
        monthly = df.copy()
        monthly["Mes"] = monthly["Data"].dt.to_period("M").dt.to_timestamp()
        monthly_total = monthly.groupby("Mes")["Receita"].sum().reset_index()
        # Write monthly_total to temporary area to plot chart
        tmp_row = chart_start_row
        ws.cell(row=tmp_row, column=1, value="Mes").font = Font(bold=True)
        ws.cell(row=tmp_row, column=2, value="Receita").font = Font(bold=True)
        for idx, row in monthly_total.iterrows():
            ws.cell(row=tmp_row + idx + 1, column=1, value=row["Mes"].strftime("%Y-%m"))
            cell = ws.cell(row=tmp_row + idx + 1, column=2, value=row["Receita"])
            cell.number_format = currency_fmt
        line_chart = LineChart()
        line_chart.title = "Receita Mensal"
        line_chart.y_axis.title = "Receita"
        line_chart.x_axis.title = "Mês"
        data_ref = Reference(ws, min_col=2, min_row=tmp_row, max_row=tmp_row + len(monthly_total))
        cats_ref = Reference(ws, min_col=1, min_row=tmp_row + 1, max_row=tmp_row + len(monthly_total))
        line_chart.add_data(data_ref, titles_from_data=True)
        line_chart.set_categories(cats_ref)
        ws.add_chart(line_chart, f"E{tmp_row}")
        # Bar chart for operadoras
        oper = df.groupby("Operadora")["Receita"].sum().reset_index()
        bar_start_row = tmp_row + len(monthly_total) + 3
        ws.cell(row=bar_start_row, column=1, value="Operadora").font = Font(bold=True)
        ws.cell(row=bar_start_row, column=2, value="Receita").font = Font(bold=True)
        for idx, row in oper.iterrows():
            ws.cell(row=bar_start_row + idx + 1, column=1, value=row["Operadora"])
            cell = ws.cell(row=bar_start_row + idx + 1, column=2, value=row["Receita"])
            cell.number_format = currency_fmt
        bar_chart = BarChart()
        bar_chart.title = "Receita por Operadora"
        bar_chart.y_axis.title = "Receita"
        bar_chart.x_axis.title = "Operadora"
        data_ref = Reference(ws, min_col=2, min_row=bar_start_row, max_row=bar_start_row + len(oper))
        cats_ref = Reference(ws, min_col=1, min_row=bar_start_row + 1, max_row=bar_start_row + len(oper))
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        ws.add_chart(bar_chart, f"E{bar_start_row}")
        # Pie chart for procedimentos
        proc = df.groupby("Procedimento")["Receita"].sum().reset_index()
        pie_start_row = bar_start_row + len(oper) + 3
        ws.cell(row=pie_start_row, column=1, value="Procedimento").font = Font(bold=True)
        ws.cell(row=pie_start_row, column=2, value="Receita").font = Font(bold=True)
        for idx, row in proc.iterrows():
            ws.cell(row=pie_start_row + idx + 1, column=1, value=row["Procedimento"])
            cell = ws.cell(row=pie_start_row + idx + 1, column=2, value=row["Receita"])
            cell.number_format = currency_fmt
        pie_chart = PieChart()
        pie_chart.title = "% Receita por Procedimento"
        labels = Reference(ws, min_col=1, min_row=pie_start_row + 1, max_row=pie_start_row + len(proc))
        data = Reference(ws, min_col=2, min_row=pie_start_row, max_row=pie_start_row + len(proc))
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        # Add the pie chart to the sheet (data labels are optional)
        ws.add_chart(pie_chart, f"E{pie_start_row}")


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------


def get_logger(cfg: Dict) -> logging.Logger:
    """Configure and return a logger that writes to the configured logs directory."""
    logs_dir = Path(cfg["saida"]["logs_dir"])
    # Resolve relative path to project root
    if not logs_dir.is_absolute():
        project_root = Path(__file__).resolve().parents[1]
        logs_dir = (project_root / logs_dir).resolve()
    logs_dir.mkdir(parents=True, exist_ok=True)
    log_path = logs_dir / "agent.log"
    logger = logging.getLogger("relatorio_saude_agent")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        fh = logging.FileHandler(log_path)
        formatter = logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
        )
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    return logger


def log_to_file_and_auditoria(params: Dict, counts: Dict[str, int], missing_segments: int, cfg: Dict) -> None:
    """Write a textual log entry to the log file."""
    logger = get_logger(cfg)
    logger.info(
        "Execução concluída. Período: %s a %s | Total lido: %d | Negativos corrigidos: %d | "
        "Duplicados removidos: %d | Fora do período removidos: %d | Filtro UF/Categoria removidos: %d | "
        "Total final: %d | Registros sem segmento: %d",
        params["periodo_inicio"],
        params["periodo_fim"],
        counts["total_lido"],
        counts["negativos_corrigidos"],
        counts["duplicados_removidos"],
        counts["fora_periodo_removidos"],
        counts["filtro_uf_categoria_removidos"],
        counts.get("final_total", 0),
        missing_segments,
    )


# ---------------------------------------------------------------------------
# Entry points called by VBA macros or CLI
# ---------------------------------------------------------------------------


def atualizar_tudo(config_path: Optional[str] = None) -> None:
    """End‑to‑end pipeline: import, clean, pivot, rankings, write workbook and log."""
    cfg = load_config(config_path)
    params = read_params_from_sheet_or_cfg(cfg)
    files = autodiscover_files(cfg["entrada"])
    raw_df = read_and_concat(files)
    cleaned_df, counts = clean_data(raw_df, params)
    cleaned_df, missing_segments = vlookup_segment(cleaned_df, cfg)
    pv_df = pivot_data(cleaned_df)
    oper_rank, proc_rank = generate_rankings(cleaned_df)
    # Persist counts final total for logging
    counts["final_total"] = len(cleaned_df)
    write_excel_sheets(cleaned_df, pv_df, oper_rank, proc_rank, params, cfg, counts, missing_segments)
    log_to_file_and_auditoria(params, counts, missing_segments, cfg)


def gerar_graficos(config_path: Optional[str] = None) -> None:
    """Regenerate only the charts on the Resumo sheet of an existing workbook."""
    cfg = load_config(config_path)
    out_path = Path(cfg["saida"]["relatorio_xlsm"])
    if not out_path.exists():
        raise FileNotFoundError(f"Workbook not found: {out_path}")
    wb = load_workbook(out_path)
    ws_resumo = wb["Resumo"] if "Resumo" in wb.sheetnames else None
    if ws_resumo is None:
        raise RuntimeError("Resumo sheet not found in workbook")
    # Read Dados sheet to recompute charts
    ws_dados = wb["Dados"] if "Dados" in wb.sheetnames else None
    if ws_dados is None:
        raise RuntimeError("Dados sheet not found in workbook")
    # Convert Dados to DataFrame to recalc charts
    data = ws_dados.values
    columns = next(data)
    df = pd.DataFrame(list(data), columns=columns)
    # Ensure date column is datetime and numeric
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    if "Receita" in df.columns:
        df["Receita"] = pd.to_numeric(df["Receita"], errors="coerce")
    # Remove previous charts
    ws_resumo._charts = []  # type: ignore
    # Currency format
    currency_fmt = "[R$-pt-BR] #,##0.00"
    _add_charts_to_resumo(ws_resumo, df, params={}, currency_fmt=currency_fmt)
    wb.save(out_path)


def gerar_rankings(config_path: Optional[str] = None) -> None:
    """Regenerate only the rankings sheet based on current data in Dados."""
    cfg = load_config(config_path)
    out_path = Path(cfg["saida"]["relatorio_xlsm"])
    if not out_path.exists():
        raise FileNotFoundError(f"Workbook not found: {out_path}")
    wb = load_workbook(out_path)
    ws_dados = wb["Dados"] if "Dados" in wb.sheetnames else None
    ws_rank = wb["Rankings"] if "Rankings" in wb.sheetnames else None
    if ws_dados is None or ws_rank is None:
        raise RuntimeError("Required sheets not found for ranking generation")
    # Read Dados into DataFrame
    data = ws_dados.values
    columns = next(data)
    df = pd.DataFrame(list(data), columns=columns)
    if "Receita" in df.columns:
        df["Receita"] = pd.to_numeric(df["Receita"], errors="coerce")
    oper_rank, proc_rank = generate_rankings(df)
    # Clear existing rankings sheet (except for header row 1 maybe)
    wb.remove(ws_rank)
    ws_rank = wb.create_sheet("Rankings")
    # Write oper_rank
    for r_idx, row in enumerate(dataframe_to_rows(oper_rank, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_rank.cell(row=r_idx + 1, column=c_idx, value=value)
    ws_rank.cell(row=1, column=1, value="Ranking de Operadoras").font = Font(bold=True)
    # Write proc_rank below
    start_row = len(oper_rank) + 5
    ws_rank.cell(row=start_row - 1, column=1, value="Ranking de Procedimentos").font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(proc_rank, index=False, header=True), start=0):
        for c_idx, value in enumerate(row, start=1):
            ws_rank.cell(row=start_row + r_idx, column=c_idx, value=value)
    # Auto column width
    for col in ws_rank.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_rank.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    wb.save(out_path)


# When executed directly, run atualizar_tudo
if __name__ == "__main__":
    atualizar_tudo()